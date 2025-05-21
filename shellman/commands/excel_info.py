import click
from openpyxl import load_workbook
from pathlib import Path

@click.command(help="""Prints a table with sheet name, rows and columns for a given Excel file.

Examples:
  shellman excel_info report.xlsx
  shellman excel_info data.xlsb
""")
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
def cli(file):
    file_path = Path(file)

    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as e:
        raise click.ClickException(f"Failed to read workbook: {e}")

    click.echo(f"{'Sheet':<20} {'Rows':>7} {'Cols':>6}")
    click.echo("-" * 34)
    for ws in wb.worksheets:
        click.echo(f"{ws.title[:20]:<20} {ws.max_row:>7} {ws.max_column:>6}")
