import click
import csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

@click.command(help="""Preview rows and columns from an Excel sheet.

Examples:
  shellman excel_preview file.xlsx --rows 10 --columns A,B --sheet Sheet1
  shellman excel_preview file.xlsx --output preview.csv
""")
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheet", default="1", help="Sheet number or name [default: 1]")
@click.option("--rows", default=20, type=int, help="Number of rows to preview")
@click.option("--columns", help="Column letters (e.g. A,C-E)")
@click.option("--output", type=click.Path(), help="Save result to CSV")
@click.option("--interactive", is_flag=True, default=False, help="Pipe output to less -S")
def cli(file, sheet, rows, columns, output, interactive):
    path = Path(file)
    wb = load_workbook(filename=path, data_only=True, read_only=True)

    try:
        if sheet.isdigit():
            ws = wb.worksheets[int(sheet) - 1]
        else:
            ws = wb[sheet]
    except Exception as e:
        raise click.ClickException(f"Invalid sheet reference: {sheet}")

    def parse_col_spec(spec):
        result = set()
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                for i in range(column_index_from_string(a), column_index_from_string(b) + 1):
                    result.add(i)
            else:
                result.add(column_index_from_string(part))
        return sorted(result)

    col_indices = parse_col_spec(columns) if columns else None

    preview = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > rows:
            break
        if col_indices:
            preview.append([row[j - 1] if j - 1 < len(row) else "" for j in col_indices])
        else:
            preview.append(list(row))

    def format_preview(rows):
        return "\n".join([",".join(str(cell) if cell is not None else "" for cell in line) for line in rows])

    output_text = format_preview(preview)

    if output:
        with open(output, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(preview)
        click.echo(f"Saved to {output}")
        if interactive:
            click.echo_via_pager(output_text)
    else:
        if interactive:
            click.echo_via_pager(output_text)
        else:
            click.echo(output_text)
            