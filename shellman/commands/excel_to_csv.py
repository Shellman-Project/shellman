import click
import csv
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from datetime import datetime

@click.command(help="""Exports sheets or parts of an Excel file to CSV.

Examples:
  shellman excel_to_csv file.xlsx --sheets Sheet1 --rows 2-10 --columns A,B --out ./csv
""")
@click.argument("file", type=click.Path(exists=True, dir_okay=False))
@click.option("--sheets", multiple=True, help="Sheet names or indexes (comma separated)")
@click.option("--rows", help="Row range in format start-end")
@click.option("--columns", help="Column letters (e.g. A,B-D)")
@click.option("--out", "out_dir", type=click.Path(), default="csv", help="Output directory")
@click.option("--overwrite", is_flag=True, help="Overwrite CSVs without timestamp")
def cli(file, sheets, rows, columns, out_dir, overwrite):
    file_path = Path(file)
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)

    def parse_column_spec(spec):
        result = set()
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                for i in range(column_index_from_string(a), column_index_from_string(b) + 1):
                    result.add(i)
            else:
                result.add(column_index_from_string(part))
        return sorted(result)

    row_start, row_end = None, None
    if rows:
        try:
            row_start, row_end = map(int, rows.split("-"))
        except Exception:
            raise click.ClickException("--rows must be in format start-end")

    col_indices = None
    if columns:
        try:
            col_indices = parse_column_spec(columns)
        except Exception:
            raise click.ClickException("--columns must be valid letters like A,B-D")

    targets = wb.sheetnames if not sheets else list(sheets)

    for name in targets:
        if name not in wb.sheetnames:
            click.echo(f"Skipping unknown sheet: {name}")
            continue
        ws = wb[name]

        csv_name = f"{name}.csv" if overwrite else f"{name}_{stamp}.csv"
        csv_path = out_path / csv_name

        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_start and i < row_start:
                    continue
                if row_end and i > row_end:
                    break
                if col_indices:
                    filtered = [row[j - 1] if j - 1 < len(row) else "" for j in col_indices]
                    writer.writerow(filtered)
                else:
                    writer.writerow(row)

        click.echo(f"Created: {csv_path}")
        