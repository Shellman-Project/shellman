# shellman/commands/excel.py
from pathlib import Path
from datetime import datetime
import csv
import importlib.resources

import click
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


# -------- helper ----------------------------------------------------------- #
def _print_help_md(sub: str, lang: str = "eng"):
    """
    Load Markdown help for given subcommand ('info' | 'preview' | 'export')
    and language ('eng' | 'pl').
    """
    lang_file = f"help_{lang.lower()}.md"
    try:
        help_path = importlib.resources.files("shellman").joinpath(
            f"help_texts/excel/{sub}/{lang_file}"
        )
        click.echo(help_path.read_text(encoding="utf-8"))
    except Exception:
        click.echo(
            f"⚠️ Help not available for '{sub}' in language '{lang}'.", err=True
        )


# -------- main click group ------------------------------------------------- #
@click.group(help="Excel utilities: sheet info, quick preview and CSV export.")
def cli():
    pass


# ======================  info  ============================================ #
@cli.command("info", help="Show all sheets with rows/columns count.")
@click.argument("file", required=False)
@click.option(
    "--lang-help",
    "lang",
    help="Show localized help (pl, eng) instead of executing the command",
)
def info(file, lang):
    if lang:
        _print_help_md("info", lang)
        return

    if not file:
        raise click.UsageError("Missing required argument 'file'")

    file_path = Path(file)
    try:
        wb = load_workbook(filename=file_path, read_only=True, data_only=True)
    except Exception as e:
        raise click.ClickException(f"Failed to read workbook: {e}")

    click.echo(f"{'Sheet':<20} {'Rows':>7} {'Cols':>6}")
    click.echo("-" * 34)
    for ws in wb.worksheets:
        click.echo(f"{ws.title[:20]:<20} {ws.max_row:>7} {ws.max_column:>6}")


# ======================  preview  ========================================= #
@cli.command("preview", help="Preview first N rows / selected columns.")
@click.argument("file", required=False)
@click.option("--sheet", default="1", help="Sheet number or name [default: 1]")
@click.option("--rows", default=20, type=int, help="Number of rows to preview")
@click.option("--columns", help="Column letters (e.g. A,C-E)")
@click.option("--output", type=click.Path(), help="Save result to CSV")
@click.option(
    "--interactive",
    is_flag=True,
    default=False,
    help="Pipe result to less -S",
)
@click.option(
    "--info",
    "-i",
    is_flag=True,
    help="Show column header + row numbers and use '|' separator",
)
@click.option("--lang-help", "lang", help="Show localized help (pl, eng)")
def preview(file, sheet, rows, columns, output, interactive, info, lang):
    if lang:
        _print_help_md("preview", lang)
        return

    if not file:
        raise click.UsageError("Missing required argument 'file'")

    path = Path(file)
    wb = load_workbook(filename=path, data_only=True, read_only=True)

    try:
        ws = wb.worksheets[int(sheet) - 1] if sheet.isdigit() else wb[sheet]
    except Exception:
        raise click.ClickException(f"Invalid sheet reference: {sheet}")

    # parse column letters like A,C-E
    def parse_spec(spec: str):
        res = set()
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                res.update(
                    range(
                        column_index_from_string(a),
                        column_index_from_string(b) + 1,
                    )
                )
            else:
                res.add(column_index_from_string(part))
        return sorted(res)

    col_indices = parse_spec(columns) if columns else None

    # ------- collect rows --------------------------------------------------
    preview_rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > rows:
            break
        if col_indices:
            preview_rows.append(
                [row[j - 1] if j - 1 < len(row) else "" for j in col_indices]
            )
        else:
            preview_rows.append(list(row))

    # ------- build output --------------------------------------------------
    if info:
        if not col_indices:
            col_indices = list(
                range(1, max(len(r) for r in preview_rows) + 1)
            )

        # convert to strings & compute width
        str_rows = [
            ["" if c is None else str(c) for c in r] for r in preview_rows
        ]
        widths = [
            max(len(str_rows[r][idx]) if idx < len(str_rows[r]) else 0 for r in range(len(str_rows)))
            for idx in range(len(col_indices))
        ]
        widths = [max(w, len(chr(64 + col))) for w, col in zip(widths, col_indices)]

        header = "Cols: " + " | ".join(
            f"{chr(64 + col):<{w}}" for col, w in zip(col_indices, widths)
        )
        out_lines = [header]

        for ridx, row in enumerate(str_rows, start=1):
            padded = [
                f"{row[idx]:<{widths[cidx]}}"
                if idx < len(row)
                else " " * widths[cidx]
                for cidx, idx in enumerate(range(len(col_indices)))
            ]
            out_lines.append(f"{ridx:>4}: " + " | ".join(padded))

        output_text = "\n".join(out_lines)
    else:
        output_text = "\n".join(
            [
                ",".join("" if cell is None else str(cell) for cell in r)
                for r in preview_rows
            ]
        )

    # ------- output / save / pager ----------------------------------------
    if output:
        Path(output).write_text(output_text, encoding="utf-8")
        click.echo(f"Saved to {output}")
        if interactive:
            click.echo_via_pager(output_text)
    else:
        if interactive:
            click.echo_via_pager(output_text)
        else:
            click.echo(output_text)


# ======================  export  ========================================== #
@cli.command("export", help="Export sheets or ranges of an Excel file to CSV.")
@click.argument("file", required=False)
@click.option("--sheets", multiple=True, help="Sheet names or indexes")
@click.option("--rows", help="Row range start-end")
@click.option("--columns", help="Column letters (e.g. A,B-D)")
@click.option(
    "--out",
    "out_dir",
    type=click.Path(),
    default="csv",
    help="Output directory",
)
@click.option("--overwrite", is_flag=True, help="Overwrite CSVs without timestamp")
@click.option("--lang-help", "lang", help="Show localized help (pl, eng)")
def export(file, sheets, rows, columns, out_dir, overwrite, lang):
    if lang:
        _print_help_md("export", lang)
        return

    if not file:
        raise click.UsageError("Missing required argument 'file'")

    file_path = Path(file)
    out_path = Path(out_dir)
    out_path.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    wb = load_workbook(filename=file_path, data_only=True, read_only=True)

    # parse column spec
    def parse_col_spec(spec: str):
        res = set()
        for part in spec.split(","):
            if "-" in part:
                a, b = part.split("-")
                res.update(
                    range(
                        column_index_from_string(a),
                        column_index_from_string(b) + 1,
                    )
                )
            else:
                res.add(column_index_from_string(part))
        return sorted(res)

    row_start, row_end = None, None
    if rows:
        try:
            row_start, row_end = map(int, rows.split("-"))
        except Exception:
            raise click.ClickException("--rows must be in format start-end")

    col_indices = None
    if columns:
        try:
            col_indices = parse_col_spec(columns)
        except Exception:
            raise click.ClickException("--columns must be like A,B-D")

    targets = wb.sheetnames if not sheets else list(sheets)

    for name in targets:
        if name not in wb.sheetnames:
            click.echo(f"Skipping unknown sheet: {name}")
            continue
        ws = wb[name]

        csv_name = f"{name}.csv" if overwrite else f"{name}_{stamp}.csv"
        csv_path = out_path / csv_name

        with csv_path.open("w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                if row_start and i < row_start:
                    continue
                if row_end and i > row_end:
                    break
                if col_indices:
                    filtered = [
                        row[j - 1] if j - 1 < len(row) else ""
                        for j in col_indices
                    ]
                    writer.writerow(filtered)
                else:
                    writer.writerow(row)

        click.echo(f"Created: {csv_path}")
